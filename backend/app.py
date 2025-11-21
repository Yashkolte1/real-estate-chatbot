# backend/app.py
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from openpyxl import load_workbook
import io, os, csv, math, datetime
from decimal import Decimal

app = Flask(__name__)
CORS(app)


def normalize_value(v):
    """Return a JSON-serializable primitive for value v."""
    if v is None:
        return None
    # keep numeric types
    if isinstance(v, (int, float, bool)):
        return v
    if isinstance(v, Decimal):
        try:
            return float(v)
        except:
            return str(v)
    # dates/times -> ISO
    if isinstance(v, (datetime.date, datetime.datetime, datetime.time)):
        try:
            return v.isoformat()
        except:
            return str(v)
    s = safe_str(v)

    s2 = s.replace(",", "").replace("₹", "").replace("$", "").replace("%", "").strip()
    if s2 == "":
        return s
    try:
        if "." in s2:
            return float(s2)
        return int(s2)
    except:
        return s


# Path to the sample file you uploaded earlier (must exist)
SAMPLE_PATH = os.path.join(os.path.dirname(__file__), "sample_data", "sample_realestate.xlsx")

# In-memory uploaded dataset (replaces sample when present)
UPLOADS = {}


def safe_str(v):
    if v is None:
        return ""
    try:
        return str(v).strip()
    except:
        return ""

def is_number(v):
    try:
        if v is None:
            return False
        # handle existing numeric types
        if isinstance(v, (int, float)):
            return True
        s = safe_str(v)
        if s == "":
            return False
        s2 = s.replace(",", "").replace("₹", "").replace("$", "").replace("%", "").strip()
        float(s2)
        return True
    except:
        return False


def to_number(v):
    """Convert v to float when possible. Return float or raise ValueError."""
    if v is None:
        raise ValueError("None")
    if isinstance(v, (int, float)):
        return float(v)
    s = safe_str(v)
    s2 = s.replace(",", "").replace("₹", "").replace("$", "").replace("%", "").strip()
    return float(s2)

def canonical_key(h):
    h = safe_str(h).lower()
    # remove punctuation and extra spaces
    h = h.replace("\n", " ").replace("\r", " ").strip()
    return h

def read_excel(path_or_file):
    """
    Read all sheets, map headers (lowercased), return list of rows as dicts.
    Each dict keys are canonical header strings (original header lowercased).
    """
    if hasattr(path_or_file, "read"):
        # uploaded file object
        data = path_or_file.read()
        wb = load_workbook(io.BytesIO(data), data_only=True)
    else:
        wb = load_workbook(path_or_file, data_only=True)

    all_rows = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.rows)
        if not rows:
            continue

        # build headers 
        raw_headers = [safe_str(c.value) for c in rows[0]]
        headers = [canonical_key(h) if h else f"col_{i}" for i, h in enumerate(raw_headers)]

        for row in rows[1:]:
            entry = {}
            raw_entry = {} 
            for idx, cell in enumerate(row):
                key = headers[idx] if idx < len(headers) else f"col_{idx}"
                val = cell.value
                norm = normalize_value(val)
                entry[key] = norm
                raw_entry[key] = norm
            # try to infer year (if not present) by checking any date-like column
            if not any(k for k in entry.keys() if k == "year"):
                for k, v in entry.items():
                    if v is None:
                        continue
                    s = safe_str(v)
                    if len(s) >= 4 and any(ch.isdigit() for ch in s):
                        digits = "".join(ch for ch in s if ch.isdigit())
                        if len(digits) >= 4:
                            try:
                                y = int(digits[:4])
                                entry["year"] = y
                                raw_entry["year"] = y
                                break
                            except:
                                pass
            all_rows.append(raw_entry)
    return all_rows


def detect_numeric_columns(rows):
    """
    Inspect first N rows and return lists: price_like_cols, demand_like_cols
    """
    if not rows:
        return [], []
    sample = rows[:min(50, len(rows))]
    cols = list(rows[0].keys())

    price_candidates = []
    demand_candidates = []

    for c in cols:
        lk = c.lower()
        if any(token in lk for token in ["price", "cost", "amount", "rate", "value", "total_sales", "sales", "total"]):
            price_candidates.append(c)
        if any(token in lk for token in ["demand", "interest", "search", "vol", "queries"]):
            demand_candidates.append(c)

   
    def keep_numeric(candidates):
        kept = []
        for c in candidates:
            numeric_count = sum(1 for r in sample if is_number(r.get(c)))
            if numeric_count >= 1:  
                kept.append(c)
        return kept

    price_cols = keep_numeric(price_candidates)
    demand_cols = keep_numeric(demand_candidates)

   
    if not price_cols:
        for c in cols:
            numeric_count = sum(1 for r in sample if is_number(r.get(c)))
            if numeric_count >= 3:
                price_cols.append(c)
                break

    return price_cols, demand_cols


def detect_area_values(rows, max_values=50):
    """Return a list of common area/locality strings found in rows."""
    if not rows:
        return []
    
    tokens = ["area", "location", "locality", "city", "final location", "neighbour", "neighborhood", "town"]
    cols = list(rows[0].keys())
    candidate_cols = [c for c in cols if any(tok in c.lower() for tok in tokens)]
    values = {}
    
    if not candidate_cols:
        candidate_cols = cols

    for r in rows:
        for c in candidate_cols:
            v = r.get(c)
            if v is None:
                continue
            s = safe_str(v)
            if not s:
                continue
            s2 = s.strip()
            
            scheck = s2.replace(',', '').replace('.', '').replace(' ', '')
            if scheck.isdigit():
                continue
            
            if len(s2) > 120:
                continue
            values[s2] = values.get(s2, 0) + 1

    
    items = sorted(values.items(), key=lambda x: x[1], reverse=True)
    return [k for k, _ in items[:max_values]]


def extract_area_from_query(q):
    if not q:
        return ""
    s = q.lower()
   
    tokens = ["analyze", "analyse", "show", "compare", "compare:", "compare ", "give", "give me analysis of", "search", "show me", "find"]
    for t in tokens:
        s = s.replace(t, "")
    return s.strip()


def filter_rows_by_area(rows, area):
    if not area:
        return rows
    a = safe_str(area).lower().strip()
    matched = []
    for r in rows:
        found = False
        for k, v in r.items():
            if v is None:
                continue
            
            vs = safe_str(v).lower()
            if a in vs:
                found = True
                break
        if found:
            matched.append(r)
    return matched


def build_time_series(rows, numeric_col):
   
    out = {}
    for r in rows:
        
        y = None
        if "year" in r and is_number(r.get("year")):
            try:
                y = int(float(r.get("year")))
            except:
                y = None
        else:
           
            for k, v in r.items():
                if v is None:
                    continue
                s = safe_str(v)
                digits = "".join(ch for ch in s if ch.isdigit())
                if len(digits) >= 4:
                    try:
                        y = int(digits[:4])
                        break
                    except:
                        pass
        if not y:
            continue
        val = r.get(numeric_col)
        if val is None:
            
            for k2, v2 in r.items():
                if is_number(v2):
                    val = float(v2)
                    break
            if val is None:
                continue
        try:
            fv = float(val)
        except:
            continue
        out.setdefault(y, []).append(fv)

    series = []
    for y in sorted(out.keys()):
        vals = out[y]
        avg = sum(vals) / len(vals) if vals else 0
        series.append({"year": y, "value": avg})
    return series


# Summary text

def build_summary(rows, area, price_col, demand_col):
    if not rows:
        return f"No records found for '{area}'. Try another locality or upload a different dataset."
    cnt = len(rows)
    years = []
    for r in rows:
        if r.get("year") and is_number(r.get("year")):
            try:
                years.append(int(float(r.get("year"))))
            except:
                pass
    yr_range = f"{min(years)}–{max(years)}" if years else "N/A"

    avg_price = None
    if price_col:
        vals = [float(r.get(price_col)) for r in rows if is_number(r.get(price_col))]
        if vals:
            avg_price = sum(vals) / len(vals)

    
    avg_demand = None
    if demand_col:
        dvals = [float(r.get(demand_col)) for r in rows if is_number(r.get(demand_col))]
        if dvals:
            avg_demand = sum(dvals) / len(dvals)

    parts = []
    parts.append(f"{cnt} records")
    parts.append(f"years: {yr_range}")
    if avg_price is not None:
        parts.append(f"avg price ≈ ₹{int(avg_price):,}")
    if avg_demand is not None:
        parts.append(f"avg demand ≈ {int(avg_demand)}")

    return f"Analysis for {area or 'All areas'}: " + ", ".join(parts) + ". Recent trend: moderate."


@app.route("/api/query", methods=["GET"])
def api_query():
    try:
        q = request.args.get("q", "")
        explicit_area = request.args.get("area", "")
        use_sample = request.args.get("use_sample", "true").lower() == "true"
        metric = request.args.get("metric", "price")  # unused currently, kept for extensibility

        # choose dataset: uploaded or sample file
        if "uploaded" in UPLOADS:
            rows = UPLOADS["uploaded"]
        elif use_sample:
           
            rows = read_excel(SAMPLE_PATH)
        else:
            return jsonify({"error": "No data available. Upload file or set use_sample=true."}), 400

        # detect price/demand columns
        price_cols, demand_cols = detect_numeric_columns(rows)
       
        price_col = request.args.get('price_col') or (price_cols[0] if price_cols else None)
        demand_col = request.args.get('demand_col') or (demand_cols[0] if demand_cols else None)

     
        if price_col and price_col not in (rows[0].keys() if rows else []):
       
            price_col = price_cols[0] if price_cols else None
        if demand_col and demand_col not in (rows[0].keys() if rows else []):
            demand_col = demand_cols[0] if demand_cols else None


        global_price_series = build_time_series(rows, price_col) if price_col else []

 
        area = explicit_area if explicit_area else extract_area_from_query(q)


        filtered = filter_rows_by_area(rows, area) if area else rows


        if filtered:
            price_series = build_time_series(filtered, price_col) if price_col else []
            demand_series = build_time_series(filtered, demand_col) if demand_col else []
        else:
            price_series = []  
            demand_series = []

        summary = build_summary(filtered, area, price_col, demand_col)
        
        table = filtered[:500]

        return jsonify({
            "summary": summary,
            "chart": price_series or global_price_series,
            "demand_chart": demand_series,   # extra field for demand
            "table": table,
            "area": area
        })
    except Exception as e:
        # return useful error message for frontend debugging and log
        import traceback
        tb = traceback.format_exc()
        print("Error in /api/query:", e)
        print(tb)
        return jsonify({"error": str(e), "trace": tb}), 500


@app.route("/api/areas", methods=["GET"])
def api_areas():
    use_sample = request.args.get("use_sample", "true").lower() == "true"
    if "uploaded" in UPLOADS:
        rows = UPLOADS["uploaded"]
    elif use_sample:
        rows = read_excel(SAMPLE_PATH)
    else:
        return jsonify({"areas": []})

    areas = detect_area_values(rows, max_values=100)
    return jsonify({"areas": areas})


@app.route("/api/columns", methods=["GET"])
def api_columns():
    """Return detected numeric columns (price-like and demand-like) from dataset."""
    use_sample = request.args.get("use_sample", "true").lower() == "true"
    if "uploaded" in UPLOADS:
        rows = UPLOADS["uploaded"]
    elif use_sample:
        rows = read_excel(SAMPLE_PATH)
    else:
        return jsonify({"price_cols": [], "demand_cols": []})

    price_cols, demand_cols = detect_numeric_columns(rows)
    return jsonify({"price_cols": price_cols, "demand_cols": demand_cols})


@app.route("/api/debug", methods=["GET"])
def api_debug():
    """Return quick diagnostics about sample file and dataset parsing."""
    info = {}
    info['sample_path'] = SAMPLE_PATH
    info['sample_exists'] = os.path.exists(SAMPLE_PATH)
    info['uploads_present'] = 'uploaded' in UPLOADS
    try:
      
        if os.path.exists(SAMPLE_PATH):
            rows = read_excel(SAMPLE_PATH)
            info['sample_rows'] = len(rows)
            info['sample_preview'] = rows[:3]
            # headers
            info['sample_headers'] = list(rows[0].keys()) if rows else []
            # detected numeric columns
            price_cols, demand_cols = detect_numeric_columns(rows)
            info['detected_price_cols'] = price_cols
            info['detected_demand_cols'] = demand_cols
          
            info['detected_areas'] = detect_area_values(rows, max_values=20)
        else:
            info['sample_rows'] = 0
            info['sample_preview'] = []
            info['sample_headers'] = []
            info['detected_price_cols'] = []
            info['detected_demand_cols'] = []
            info['detected_areas'] = []
    except Exception as e:
        # very defensive: return exception info but avoid raising
        import traceback
        info['error_reading_sample'] = str(e)
        info['trace'] = traceback.format_exc()


    if 'uploaded' in UPLOADS:
        info['uploads_rows'] = len(UPLOADS['uploaded'])
        info['uploads_preview'] = UPLOADS['uploaded'][:3]
    return jsonify(info)

@app.route("/api/upload", methods=["POST"])
def api_upload():
    if "file" not in request.files:
        return jsonify({"error": "No file provided."}), 400
    f = request.files["file"]
    try:
        data_rows = read_excel(f)
    except Exception as e:
        return jsonify({"error": f"Failed to parse uploaded file: {e}"}), 500
    UPLOADS["uploaded"] = data_rows
    return jsonify({"message": "File uploaded and parsed successfully.", "rows": len(data_rows)})

@app.route("/api/download", methods=["GET"])
def api_download():
    area = request.args.get("area", "")
    use_sample = request.args.get("use_sample", "true").lower() == "true"

    if "uploaded" in UPLOADS:
        rows = UPLOADS["uploaded"]
    elif use_sample:
        rows = read_excel(SAMPLE_PATH)
    else:
        return jsonify({"error": "No data available."}), 400

    filtered = filter_rows_by_area(rows, area) if area else rows
    if not filtered:
        return jsonify({"error": "No rows found."}), 400

    # use keys from first row for CSV headers
    keys = list(filtered[0].keys())
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(keys)
    for r in filtered:
        writer.writerow([r.get(k, "") for k in keys])
    out.seek(0)
    return send_file(io.BytesIO(out.getvalue().encode("utf-8")), mimetype="text/csv", as_attachment=True, download_name=f"filtered_{area or 'all'}.csv")


@app.route("/", methods=["GET"])
def home():
    return "Backend running"

if __name__ == "__main__":
    
    if not os.path.exists(SAMPLE_PATH):
        print("WARNING: sample file not found at", SAMPLE_PATH)
    app.run(debug=True, port=8000)
